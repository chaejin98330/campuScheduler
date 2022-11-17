import openpyxl
import math
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

def printExcel(result, student_list):
    wb = openpyxl.Workbook()
    sheet = wb.active
    grey_color = PatternFill(start_color='e6e6e6', end_color='e6e6e6', fill_type='solid')
    sheet.column_dimensions['A'].width = 12
    for col in range(2,7):
        sheet.column_dimensions[get_column_letter(col)].width = 15
        sheet.cell(1, col).fill = grey_color
    for row in range(2,28):
        sheet.row_dimensions[row].height = 25

    time = 9
    j = 2
    for i in range(13):
        sheet.cell(row=j, column=1).value = str(time) + ":00~" + str(time) + ":30"
        sheet.cell(row=j+1, column=1).value = str(time) + ":30~" + str(time+1) + ":00"
        sheet.cell(row=j, column=1).fill = grey_color
        sheet.cell(row=j+1, column=1).fill = grey_color
        time += 1
        j += 2
    
    sheet['B1'] = "Mon"
    sheet['C1'] = "Tue"
    sheet['D1'] = "Wed"
    sheet['E1'] = "Thu"
    sheet['F1'] = "Fri"

    for i in range(len(result)):
        name = student_list[i].name
        id = student_list[i].sid
        for j in range(len(result[i])):
            start = result[i][j][0]
            end = result[i][j][1]
            day = math.floor(start / (24*60))
            start = (start - day * 24 * 60)/60
            end = (end - day * 24 * 60)/60
            
            start_r = (math.floor(start) - 8) * 2
            if not float(start).is_integer():
                start_r += 1
            end_r = (math.floor(end) - 8) * 2 - 1
            if not float(end).is_integer():
                end_r += 1

            sheet.merge_cells(start_row=start_r, start_column=day+2, end_row=end_r, end_column=day+2)
            if sheet.cell(row=start_r, column=day+2).value == None:
                sheet.cell(row=start_r, column=day+2).value = name + "(" + str(id) + ")"
            else:
                sheet.cell(row=start_r, column=day+2).value += ",\n" + name + "(" + str(id) + ")"

    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save('test.xlsx')